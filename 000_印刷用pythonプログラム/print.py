from websocket_server import WebsocketServer
import win32api
import sys
import os
import time
import openpyxl
from openpyxl.styles.borders import Border, Side
import json
import shutil
import datetime

repeat = []
#プリント実行関数
def auto_print(path):
    print(path)
    return_value = win32api.ShellExecute(0,"print",path,None,".",0)
    print("result_code" + str(return_value))
    print("printed:" + path)

#印刷用フォルダ中のファイルを検索
def file_check(path):
    if os.path.isdir(path):
        files = os.listdir(path)
        for file in files:
            file_check(path+"\\"+file)
    else:
        for i in range(repeat[0]):
            auto_print(path)
            time.sleep(3)
        del repeat[0]

def write_excel(dict):
    cnt = 0
    template_wb = openpyxl.load_workbook(r"./template/template.xlsx")
    header = dict["nyuko_header"]
    del dict["nyuko_header"]
    for key, rec in dict.items():
        template_wb.save(r"./print/print" + str(cnt) + ".xlsx")
        print_wb = openpyxl.load_workbook(r"./print/print" + str(cnt) + ".xlsx")
        ws = print_wb.worksheets[0]
        C2 = ws["C2"]
        C3 = ws["C3"]
        E5 = ws["E5"]
        E6 = ws["E6"]
        E7 = ws["E7"]
        E8 = ws["E8"]
        E9 = ws["E9"]
        C2.value = header["date"]
        C3.value = header["staff_ID"]
        E5.value = rec["code"]
        E6.value = rec["wakuchin_name"]
        E7.value = rec["expair"]
        E8.value = rec["lot"]
        if rec["wakuchin_name"] == "コミナティ筋注":
            E9.value = '195'
            repeat.append(int(rec["amount"]/195))
        elif rec["wakuchin_name"] == "バキスゼブリア筋注":
            E9.value = 2
            repeat.append(int(rec["amount"]/2))
        else:
            E9.value = 10
            repeat.append(int(rec["amount"]/10))
        ws.print_area = "A1:H39"
        ws.oddHeader.center.text = "ワクチン管理帳票"
        wps = ws.page_setup
        wps.paperSize = ws.PAPERSIZE_A4
        print_wb.save(r"./print/print" + str(cnt) + ".xlsx")
        cnt += 1

def write_excel2(dict):
    currentRow = 5
    template_wb = openpyxl.load_workbook(r"./template/template2.xlsx")
    template_wb.save(r"./print/print.xlsx")
    print_wb = openpyxl.load_workbook(r"./print/print.xlsx")
    ws = print_wb.worksheets[0] 
    template_wb.save(r"./print/print.xlsx")
    side1 = Side(style='thin', color='000000')
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    date = ws["F2"]
    print(datetime.date.today())
    date.value = datetime.date.today()
    for key, rec in dict.items():
        code = ws["B" + str(currentRow)]  
        name = ws["D" + str(currentRow)]
        lot = ws["F" + str(currentRow)]
        quantity = ws["G" + str(currentRow)]
        code.value = rec['wakuchin_ID']
        name.value = rec['wakuchin_name']
        lot.value = rec['lot_number']
        quantity.value = rec['yotei_amount']
        lineAddress = "B" + str(currentRow) + ":" + "G" + str(currentRow)
        for row in ws[lineAddress]:
            for cell in row:
                cell.border = border_aro
        currentRow += 1
        repeat.append(1)
    wps = ws.page_setup
    wps.paperSize = ws.PAPERSIZE_A4
    ws.print_area = "A1:H39"
    print_wb.save(r"./print/print.xlsx")
    
def write_excel3(dict):
    cnt = 0
    template_wb = openpyxl.load_workbook(r"./template/template3.xlsx")
    header = dict["header"]
    del dict["header"]
    for key, rec in dict.items():
        template_wb.save(r"./print/print" + str(cnt) + ".xlsx")
        print_wb = openpyxl.load_workbook(r"./print/print" + str(cnt) + ".xlsx")
        ws = print_wb.worksheets[0]
        C2 = ws["C2"]
        C3 = ws["C3"]
        E5 = ws["E5"]
        E6 = ws["E6"]
        E7 = ws["E7"]
        E8 = ws["E8"]
        E9 = ws["E9"]
        C2.value = header["current_date"]
        C3.value = header["staff_ID"]
        E5.value = rec["code"]
        E6.value = rec["name"]
        E7.value = rec["lot"]
        E8.value = header["next_month"]
        E9.value = rec["amount"]
        repeat.append(int(rec["print_num"]))
        ws.print_area = "A1:H39"
        ws.oddHeader.center.text = "ワクチン管理帳票"
        wps = ws.page_setup
        wps.paperSize = ws.PAPERSIZE_A4
        print_wb.save(r"./print/print" + str(cnt) + ".xlsx")
        cnt += 1

            



def receivedMessage(client, server, message):
    print_id_dict = {1:'nyuko',2:'defrost_list',3:'defrost_label'}
    print_path = r"./print"
    records = json.loads(message)
    print(records)
    print(type(records))
    if print_id_dict[records['print_id']] == 'nyuko':
        del records['print_id']
        print('入庫')
        write_excel(records)
        file_check(print_path)
        shutil.rmtree(print_path)
        os.mkdir(print_path)
        print("Process finished")
        server.send_message_to_all("complete")
    elif print_id_dict[records['print_id']] == 'defrost_list':
        del records['print_id']
        print('解凍')
        write_excel2(records)
        file_check(print_path)
        shutil.rmtree(print_path)
        os.mkdir(print_path)
        server.send_message_to_all("complete")
    elif print_id_dict[records['print_id']] == 'defrost_label':
        del records['print_id']
        print('解凍ラベル')
        write_excel3(records)
        file_check(print_path)
        shutil.rmtree(print_path)
        os.mkdir(print_path)
        server.send_message_to_all("complete")
#新規クライアント接続時の処理    
def new_client(client, server):
    print("クライアント接続")

#ウェブソケットサーバ起動メソッド
print("サーバー起動")
server = WebsocketServer(host="127.0.0.1", port=9999)
server.set_fn_new_client(new_client)
server.set_fn_message_received(receivedMessage)
server.run_forever()


 
