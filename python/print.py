from websocket_server import WebsocketServer
import win32api
import sys
import os
import time
import openpyxl
import json

repeat = []
#プリント実行関数
def auto_print(path):
    print(path)
    return_value = win32api.ShellExecute(0,"print",path,None,".",0)
    print("result_code" + str(return_value))
    print("printed:" + path)

#印刷用フォルダ中のファイルを検索
def file_check(path):
    if os.path.isdir(path):
        files = os.listdir(path)
        for file in files:
            file_check(path+"\\"+file)
    else:
        for i in range(repeat[0]):
            auto_print(path)
            time.sleep(3)
        del repeat[0]

def write_excel(dict):
    cnt = 0
    template_wb = openpyxl.load_workbook(r"./template/template.xlsx")
    header = dict["nyuko_header"]
    del dict["nyuko_header"]
    for key, rec in dict.items():
        template_wb.save(r"./print/print" + str(cnt) + ".xlsx")
        print_wb = openpyxl.load_workbook(r"./print/print" + str(cnt) + ".xlsx")
        ws = print_wb.worksheets[0]
        C2 = ws["C2"]
        C3 = ws["C3"]
        E5 = ws["E5"]
        E6 = ws["E6"]
        E7 = ws["E7"]
        E8 = ws["E8"]
        E9 = ws["E9"]
        C2.value = header["staff_ID"]
        C3.value = header["hospital_ID"]
        E5.value = rec["code"]
        E6.value = rec["wakuchin_name"]
        E7.value = rec["expair"]
        E8.value = rec["lot"]
        if rec["wakuchin_name"] == "コミナティ筋注":
            E9.value = int(rec["amount"]/195)
            repeat.append(int(rec["amount"]/195))
        elif rec["wakuchin_name"] == "バキスゼブリア筋注":
            E9.value = int(rec["amount"]/2)
            repeat.append(int(rec["amount"]/2))
        else:
            E9.value = int(rec["amount"]/10)
            repeat.append(int(rec["amount"]/10))
        ws.print_area = "A1:H39"
        ws.oddHeader.center.text = "ワクチン管理帳票"
        wps = ws.page_setup
        wps.paperSize = ws.PAPERSIZE_A4
        print_wb.save(r"./print/print" + str(cnt) + ".xlsx")
        cnt += 1
        
            
            



def receivedMessage(client, server, message):
    print_path = r"./print"
    nyuko_records = json.loads(message)
    print(nyuko_records)
    write_excel(nyuko_records)
    file_check(print_path)
    print("Process finished")
    server.send_message_to_all("complete")
    
    
def new_client(client, server):
    print("クライアント接続")
    
print("サーバー起動")
server = WebsocketServer(host="127.0.0.1", port=9999)
server.set_fn_new_client(new_client)
server.set_fn_message_received(receivedMessage)
server.run_forever()


 
